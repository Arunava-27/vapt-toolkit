#!/usr/bin/env python3
"""
VAPT Toolkit — Phase 6 Manual Testing Excel Generator
Creates comprehensive test results spreadsheet
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Install with: pip install openpyxl")

from datetime import datetime
from pathlib import Path

def create_test_results_excel():
    """Create comprehensive test results spreadsheet."""
    
    if not EXCEL_AVAILABLE:
        print("❌ openpyxl not available. Install with: pip install openpyxl")
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    category_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    category_font = Font(bold=True, size=10)
    
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    blocked_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    pending_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Test ID", "Test Name", "Category", "Environment", 
        "Description", "Status", "Tester", "Date", "Expected Result", 
        "Actual Result", "Pass/Fail", "Notes", "Screenshot"
    ]
    
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    
    # Test data
    test_data = [
        # Web Scanning
        ("WS-001", "SQL Injection Detection", "Web Scanning", "DVWA+Juice Shop",
         "Detect SQL injection in login forms", "", "", "", "SQLi detected, severity HIGH",
         "", "", "", ""),
        ("WS-002", "Reflected XSS Detection", "Web Scanning", "DVWA+Juice Shop",
         "Detect reflected cross-site scripting", "", "", "", "XSS payload identified",
         "", "", "", ""),
        ("WS-003", "Stored XSS Detection", "Web Scanning", "Juice Shop+WebGoat",
         "Detect stored XSS in comments", "", "", "", "Stored XSS flagged as persistent",
         "", "", "", ""),
        ("WS-004", "CSRF Detection", "Web Scanning", "Juice Shop+WebGoat",
         "Detect CSRF token missing", "", "", "", "CSRF vulnerability documented",
         "", "", "", ""),
        ("WS-005", "IDOR Detection", "Web Scanning", "Juice Shop",
         "Detect insecure object references", "", "", "", "IDOR vulnerability identified",
         "", "", "", ""),
        ("WS-006", "Authentication Bypass", "Web Scanning", "DVWA+WebGoat",
         "Detect auth bypass vectors", "", "", "", "Auth bypass techniques documented",
         "", "", "", ""),
        ("WS-007", "Insecure Cookies", "Web Scanning", "All",
         "Check cookie security flags", "", "", "", "Missing HttpOnly/Secure flags identified",
         "", "", "", ""),
        ("WS-008", "Missing Security Headers", "Web Scanning", "All",
         "Detect missing HTTP security headers", "", "", "", "Missing headers documented",
         "", "", "", ""),
        ("WS-009", "File Upload Vulnerabilities", "Web Scanning", "Juice Shop+DVWA",
         "Test file upload restrictions", "", "", "", "Unrestricted upload detected",
         "", "", "", ""),
        ("WS-010", "Path Traversal", "Web Scanning", "DVWA+WebGoat",
         "Detect directory traversal", "", "", "", "Path traversal vulnerability found",
         "", "", "", ""),
        ("WS-011", "Command Injection", "Web Scanning", "DVWA",
         "Detect OS command injection", "", "", "", "Command injection vulnerability detected",
         "", "", "", ""),
        ("WS-012", "XXE Injection", "Web Scanning", "WebGoat",
         "Detect XML external entity injection", "", "", "", "XXE vulnerability documented",
         "", "", "", ""),
        ("WS-013", "SSRF Detection", "Web Scanning", "Juice Shop",
         "Detect server-side request forgery", "", "", "", "SSRF vulnerability identified",
         "", "", "", ""),
        ("WS-014", "Open Redirect", "Web Scanning", "Juice Shop",
         "Detect open redirect", "", "", "", "Open redirect vulnerability found",
         "", "", "", ""),
        ("WS-015", "Information Disclosure", "Web Scanning", "All",
         "Detect sensitive information leakage", "", "", "", "Information disclosure items found",
         "", "", "", ""),
        
        # API Tests
        ("API-001", "API Key Authentication", "API", "VAPT Toolkit",
         "Test API key validation", "", "", "", "Valid key accepted, invalid rejected",
         "", "", "", ""),
        ("API-002", "Rate Limiting", "API", "VAPT Toolkit",
         "Test rate limiting enforcement", "", "", "", "Rate limit enforced at 100 req/min",
         "", "", "", ""),
        ("API-003", "Bulk Scanning", "API", "VAPT Toolkit",
         "Test bulk scanning API", "", "", "", "Bulk job processes targets correctly",
         "", "", "", ""),
        ("API-004", "Export Functionality", "API", "VAPT Toolkit",
         "Test multiple export formats", "", "", "", "PDF, Excel, JSON exports working",
         "", "", "", ""),
        ("API-005", "Webhook Delivery", "API", "VAPT Toolkit",
         "Test webhook event delivery", "", "", "", "Webhook events delivered and logged",
         "", "", "", ""),
        
        # Reporting Tests
        ("REP-001", "PDF Report Generation", "Reporting", "VAPT Toolkit",
         "Test PDF report generation", "", "", "", "PDF generated with all sections",
         "", "", "", ""),
        ("REP-002", "Excel Export", "Reporting", "VAPT Toolkit",
         "Test Excel export", "", "", "", "Excel file valid with proper formatting",
         "", "", "", ""),
        ("REP-003", "Scan Comparison", "Reporting", "VAPT Toolkit",
         "Test vulnerability comparison", "", "", "", "Delta analysis accurate",
         "", "", "", ""),
        ("REP-004", "Heat Map Rendering", "Reporting", "VAPT Toolkit",
         "Test heat map visualization", "", "", "", "Heat map renders without errors",
         "", "", "", ""),
        ("REP-005", "Executive Report", "Reporting", "VAPT Toolkit",
         "Test executive summary generation", "", "", "", "Executive report complete",
         "", "", "", ""),
        
        # UX Tests
        ("UX-001", "Scope Editor Workflow", "UX", "VAPT Toolkit",
         "Test scope editor functionality", "", "", "", "Scope editor fully functional",
         "", "", "", ""),
        ("UX-002", "Theme Switching", "UX", "VAPT Toolkit",
         "Test theme switching", "", "", "", "Themes switch and persist",
         "", "", "", ""),
        ("UX-003", "Notification Delivery", "UX", "VAPT Toolkit",
         "Test real-time notifications", "", "", "", "Notifications appear and clear properly",
         "", "", "", ""),
        ("UX-004", "Schedule Creation/Execution", "UX", "VAPT Toolkit",
         "Test scan scheduling", "", "", "", "Schedules created and executed",
         "", "", "", ""),
        ("UX-005", "Search/Filter Functionality", "UX", "VAPT Toolkit",
         "Test search and filtering", "", "", "", "Search/filter results accurate",
         "", "", "", ""),
    ]
    
    # Add test rows
    for row_idx, row_data in enumerate(test_data, start=2):
        ws.append(row_data)
        
        # Apply styling
        for cell_idx, cell in enumerate(ws[row_idx], start=1):
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Status cell styling
            if cell_idx == 6:  # Status column
                if cell.value == "PASS":
                    cell.fill = pass_fill
                elif cell.value == "FAIL":
                    cell.fill = fail_fill
                elif cell.value == "BLOCKED":
                    cell.fill = blocked_fill
                else:
                    cell.fill = pending_fill
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 10
    ws.column_dimensions['L'].width = 30
    ws.column_dimensions['M'].width = 20
    
    # Add statistics sheet
    stats_sheet = wb.create_sheet("Statistics")
    
    stats_headers = ["Metric", "Value", "Percentage"]
    stats_sheet.append(stats_headers)
    for cell in stats_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    stats_data = [
        ["Total Tests", 35, "100%"],
        ["Passed", 0, "0%"],
        ["Failed", 0, "0%"],
        ["Blocked", 0, "0%"],
        ["Not Started", 35, "100%"],
        ["", "", ""],
        ["Pass Rate", "0%", ""],
        ["", "", ""],
        ["Tests by Category", "", ""],
        ["Web Scanning", 15, "43%"],
        ["API Tests", 5, "14%"],
        ["Reporting Tests", 5, "14%"],
        ["UX Tests", 5, "14%"],
        ["Edge Cases", 5, "14%"],
    ]
    
    for row_data in stats_data:
        stats_sheet.append(row_data)
    
    for row in stats_sheet.iter_rows(min_row=2, max_row=len(stats_data)+1):
        for cell in row:
            cell.border = border
    
    stats_sheet.column_dimensions['A'].width = 25
    stats_sheet.column_dimensions['B'].width = 15
    stats_sheet.column_dimensions['C'].width = 15
    
    # Add defects sheet
    defects_sheet = wb.create_sheet("Defects")
    
    defect_headers = [
        "Defect ID", "Title", "Severity", "Priority", "Status",
        "Related Test", "Description", "Steps to Reproduce", "Expected", "Actual", "Resolution"
    ]
    defects_sheet.append(defect_headers)
    for cell in defects_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Placeholder for defects
    defects_sheet.append(["(No defects found yet)", "", "", "", "", "", "", "", "", "", ""])
    
    for col_idx in range(1, len(defect_headers) + 1):
        defects_sheet.column_dimensions[get_column_letter(col_idx)].width = 20
    
    # Save
    output_file = Path("test_results") / "testing_results.xlsx"
    output_file.parent.mkdir(exist_ok=True)
    wb.save(output_file)
    
    print(f"✅ Created test results spreadsheet: {output_file}")
    print(f"   - Sheet 1: Test Results (35 test cases)")
    print(f"   - Sheet 2: Statistics")
    print(f"   - Sheet 3: Defects Log")

if __name__ == "__main__":
    create_test_results_excel()
