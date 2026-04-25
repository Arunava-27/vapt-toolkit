"""
Professional Excel Export Generator for VAPT Scan Results.
Creates multi-sheet Excel workbooks with formatting, charts, and professional styling.
"""

from __future__ import annotations
import io
from datetime import datetime
from typing import List, Dict, Any, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
        ProtectedCell, DEFAULT_FONT
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelExporter:
    """Generate professional Excel reports for scan results."""

    # Color scheme
    COLORS = {
        "critical": "FF0000",
        "high": "FF6B00",
        "medium": "FFC600",
        "low": "92D050",
        "info": "00B0F0",
        "header": "366092",
        "summary": "D9E1F2",
        "alt_row": "F2F2F2",
    }

    def __init__(self, target: str):
        """
        Initialize Excel exporter.
        
        Args:
            target: Scan target URL/IP
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            )
        self.target = target
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)

    def generate(
        self,
        findings: List[Dict[str, Any]],
        metadata: Optional[Dict[str, Any]] = None,
        include_evidence: bool = True,
    ) -> bytes:
        """
        Generate Excel workbook.
        
        Args:
            findings: List of findings
            metadata: Scan metadata
            include_evidence: Include evidence data
            
        Returns:
            Excel file bytes
        """
        # Create sheets
        self._create_summary_sheet(metadata, findings)
        self._create_findings_sheet(findings, include_evidence)
        
        if include_evidence and any(f.get("evidence") for f in findings):
            self._create_evidence_sheet(findings)

        self._create_timeline_sheet(findings)
        self._create_statistics_sheet(findings, metadata)

        # Save to bytes
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output.getvalue()

    def _get_header_style(self):
        """Get header cell style."""
        return {
            "font": Font(bold=True, color="FFFFFF", size=11),
            "fill": PatternFill(start_color=self.COLORS["header"], end_color=self.COLORS["header"], fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            ),
        }

    def _get_severity_fill(self, severity: str) -> PatternFill:
        """Get fill color based on severity."""
        color = self.COLORS.get(severity.lower(), "FFFFFF")
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def _apply_style(self, cell, **kwargs):
        """Apply style to cell."""
        if "font" in kwargs:
            cell.font = kwargs["font"]
        if "fill" in kwargs:
            cell.fill = kwargs["fill"]
        if "alignment" in kwargs:
            cell.alignment = kwargs["alignment"]
        if "border" in kwargs:
            cell.border = kwargs["border"]

    def _create_summary_sheet(self, metadata: Optional[Dict], findings: List[Dict]):
        """Create summary/overview sheet."""
        ws = self.workbook.create_sheet("Summary", 0)
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 50

        # Title
        ws["A1"] = "VAPT Scan Report - Summary"
        ws["A1"].font = Font(bold=True, size=16, color=self.COLORS["header"])
        ws.merge_cells("A1:B1")

        row = 3
        
        if metadata:
            # Metadata section
            ws[f"A{row}"] = "Scan Information"
            ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
            ws[f"A{row}"].fill = PatternFill(start_color=self.COLORS["header"], end_color=self.COLORS["header"], fill_type="solid")
            ws.merge_cells(f"A{row}:B{row}")
            row += 1

            info_items = [
                ("Target", metadata.get("target", "")),
                ("Scan Date", metadata.get("scan_date", "")),
                ("Scan Type", metadata.get("scan_type", "")),
                ("Total Findings", str(metadata.get("total_findings", 0))),
                ("Duration", metadata.get("duration", "")),
            ]

            for label, value in info_items:
                ws[f"A{row}"] = label
                ws[f"A{row}"].font = Font(bold=True)
                ws[f"B{row}"] = value
                row += 1

            row += 1

            # Severity Summary
            ws[f"A{row}"] = "Findings by Severity"
            ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
            ws[f"A{row}"].fill = PatternFill(start_color=self.COLORS["header"], end_color=self.COLORS["header"], fill_type="solid")
            ws.merge_cells(f"A{row}:B{row}")
            row += 1

            severity_counts = metadata.get("findings_by_severity", {})
            for severity in ["critical", "high", "medium", "low", "info"]:
                count = severity_counts.get(severity, 0)
                ws[f"A{row}"] = severity.capitalize()
                ws[f"A{row}"].font = Font(bold=True)
                ws[f"B{row}"] = count
                ws[f"B{row}"].fill = self._get_severity_fill(severity)
                ws[f"B{row}"].font = Font(bold=True, color="FFFFFF")
                ws[f"B{row}"].alignment = Alignment(horizontal="center")
                row += 1
        else:
            # Simple summary
            ws[f"A{row}"] = "Total Findings"
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = len(findings)
            row += 1

        row += 2

        # Finding type summary
        ws[f"A{row}"] = "Findings by Type"
        ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill(start_color=self.COLORS["header"], end_color=self.COLORS["header"], fill_type="solid")
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

        type_counts = {}
        for finding in findings:
            ftype = finding.get("type", "unknown")
            type_counts[ftype] = type_counts.get(ftype, 0) + 1

        for ftype, count in sorted(type_counts.items()):
            ws[f"A{row}"] = ftype.replace("_", " ").title()
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = count
            ws[f"B{row}"].alignment = Alignment(horizontal="center")
            row += 1

    def _create_findings_sheet(self, findings: List[Dict], include_evidence: bool):
        """Create detailed findings sheet."""
        ws = self.workbook.create_sheet("Findings")

        # Headers
        headers = [
            "ID",
            "Type",
            "Title",
            "Severity",
            "Confidence",
            "Location",
            "Description",
            "Remediation",
            "CWE ID",
            "OWASP",
        ]

        if include_evidence:
            headers.insert(7, "Evidence")

        header_style = self._get_header_style()

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            self._apply_style(cell, **header_style)

        # Set column widths
        column_widths = {
            1: 5,    # ID
            2: 15,   # Type
            3: 25,   # Title
            4: 12,   # Severity
            5: 12,   # Confidence
            6: 20,   # Location
            7: 30,   # Description
            8: 25,   # Remediation
            9: 10,   # CWE ID
            10: 20,  # OWASP
        }

        if include_evidence:
            column_widths[8] = 30  # Evidence

        for col, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        # Add findings
        for row_idx, finding in enumerate(findings, 2):
            col_idx = 1

            # ID
            ws.cell(row=row_idx, column=col_idx).value = row_idx - 1
            col_idx += 1

            # Type
            ws.cell(row=row_idx, column=col_idx).value = finding.get("type", "")
            col_idx += 1

            # Title
            ws.cell(row=row_idx, column=col_idx).value = finding.get("title", "")
            col_idx += 1

            # Severity
            severity_cell = ws.cell(row=row_idx, column=col_idx)
            severity_cell.value = finding.get("severity", "").upper()
            severity_cell.fill = self._get_severity_fill(finding.get("severity", ""))
            severity_cell.alignment = Alignment(horizontal="center")
            col_idx += 1

            # Confidence
            ws.cell(row=row_idx, column=col_idx).value = finding.get("confidence", "").capitalize()
            col_idx += 1

            # Location
            ws.cell(row=row_idx, column=col_idx).value = finding.get("location", "")
            col_idx += 1

            # Description
            desc_cell = ws.cell(row=row_idx, column=col_idx)
            desc_cell.value = finding.get("description", "")
            desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
            col_idx += 1

            # Evidence (if included)
            if include_evidence:
                evidence_cell = ws.cell(row=row_idx, column=col_idx)
                evidence_cell.value = finding.get("evidence", "")
                evidence_cell.alignment = Alignment(wrap_text=True, vertical="top")
                col_idx += 1

            # Remediation
            remediation_cell = ws.cell(row=row_idx, column=col_idx)
            remediation_cell.value = finding.get("remediation", "")
            remediation_cell.alignment = Alignment(wrap_text=True, vertical="top")
            col_idx += 1

            # CWE ID
            if finding.get("cwe_id"):
                ws.cell(row=row_idx, column=col_idx).value = f"CWE-{finding.get('cwe_id')}"
                ws.cell(row=row_idx, column=col_idx).font = Font(color="0563C1", underline="single")
            col_idx += 1

            # OWASP
            ws.cell(row=row_idx, column=col_idx).value = finding.get("owasp", "")
            col_idx += 1

        # Freeze panes
        ws.freeze_panes = "A2"

    def _create_evidence_sheet(self, findings: List[Dict]):
        """Create evidence sheet for findings with evidence."""
        ws = self.workbook.create_sheet("Evidence")
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 80

        # Headers
        for col_idx, header in enumerate(["Finding ID", "Title", "Evidence"], 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            self._apply_style(cell, **self._get_header_style())

        row = 2
        for idx, finding in enumerate(findings, 1):
            if finding.get("evidence"):
                ws.cell(row=row, column=1).value = idx
                ws.cell(row=row, column=2).value = finding.get("title", "")
                
                evidence_cell = ws.cell(row=row, column=3)
                evidence_cell.value = finding.get("evidence", "")
                evidence_cell.alignment = Alignment(wrap_text=True, vertical="top")
                
                row += 1

    def _create_timeline_sheet(self, findings: List[Dict]):
        """Create timeline sheet showing findings discovery."""
        ws = self.workbook.create_sheet("Timeline")
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 12

        # Headers
        for col_idx, header in enumerate(["Index", "Type", "Title", "Severity"], 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            self._apply_style(cell, **self._get_header_style())

        # Add findings in order
        for row_idx, finding in enumerate(findings, 2):
            ws.cell(row=row_idx, column=1).value = row_idx - 1
            ws.cell(row=row_idx, column=2).value = finding.get("type", "")
            ws.cell(row=row_idx, column=3).value = finding.get("title", "")
            
            severity_cell = ws.cell(row=row_idx, column=4)
            severity_cell.value = finding.get("severity", "").upper()
            severity_cell.fill = self._get_severity_fill(finding.get("severity", ""))
            severity_cell.alignment = Alignment(horizontal="center")

    def _create_statistics_sheet(self, findings: List[Dict], metadata: Optional[Dict]):
        """Create statistics sheet with analysis."""
        ws = self.workbook.create_sheet("Statistics")
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15

        row = 1

        # Severity Statistics
        ws[f"A{row}"] = "Severity Distribution"
        ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill(start_color=self.COLORS["header"], end_color=self.COLORS["header"], fill_type="solid")
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

        severity_counts = {}
        for finding in findings:
            severity = finding.get("severity", "info").lower()
            severity_counts[severity] = severity_counts.get(severity, 0) + 1

        for severity in ["critical", "high", "medium", "low", "info"]:
            count = severity_counts.get(severity, 0)
            ws[f"A{row}"] = severity.capitalize()
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = count
            ws[f"B{row}"].fill = self._get_severity_fill(severity)
            ws[f"B{row}"].alignment = Alignment(horizontal="center")
            row += 1

        row += 2

        # Type Statistics
        ws[f"A{row}"] = "Type Distribution"
        ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill(start_color=self.COLORS["header"], end_color=self.COLORS["header"], fill_type="solid")
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

        type_counts = {}
        for finding in findings:
            ftype = finding.get("type", "unknown")
            type_counts[ftype] = type_counts.get(ftype, 0) + 1

        for ftype, count in sorted(type_counts.items(), key=lambda x: x[1], reverse=True):
            ws[f"A{row}"] = ftype.replace("_", " ").title()
            ws[f"B{row}"] = count
            ws[f"B{row}"].alignment = Alignment(horizontal="center")
            row += 1

        row += 2

        # Confidence Statistics
        ws[f"A{row}"] = "Confidence Distribution"
        ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill(start_color=self.COLORS["header"], end_color=self.COLORS["header"], fill_type="solid")
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

        confidence_counts = {}
        for finding in findings:
            confidence = finding.get("confidence", "unknown").lower()
            confidence_counts[confidence] = confidence_counts.get(confidence, 0) + 1

        for confidence in ["high", "medium", "low"]:
            count = confidence_counts.get(confidence, 0)
            if count > 0:
                ws[f"A{row}"] = confidence.capitalize()
                ws[f"B{row}"] = count
                ws[f"B{row}"].alignment = Alignment(horizontal="center")
                row += 1

    """Generate professional Excel reports for scan results."""

    # Color scheme
    COLORS = {
        "critical": "FF0000",
        "high": "FF6B00",
        "medium": "FFC600",
        "low": "92D050",
        "info": "00B0F0",
        "header": "366092",
        "summary": "D9E1F2",
        "alt_row": "F2F2F2",
    }

    def __init__(self, target: str):
        """
        Initialize Excel exporter.
        
        Args:
            target: Scan target URL/IP
        """
        self.target = target
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)

    def generate(
        self,
        findings: List[Dict[str, Any]],
        metadata: Optional[Dict[str, Any]] = None,
        include_evidence: bool = True,
    ) -> bytes:
        """
        Generate Excel workbook.
        
        Args:
            findings: List of findings
            metadata: Scan metadata
            include_evidence: Include evidence data
            
        Returns:
            Excel file bytes
        """
        # Create sheets
        self._create_summary_sheet(metadata, findings)
        self._create_findings_sheet(findings, include_evidence)
        
        if include_evidence and any(f.get("evidence") for f in findings):
            self._create_evidence_sheet(findings)

        self._create_timeline_sheet(findings)
        self._create_statistics_sheet(findings, metadata)

        # Save to bytes
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output.getvalue()

    def _get_header_style(self):
        """Get header cell style."""
        return {
            "font": Font(bold=True, color="FFFFFF", size=11),
            "fill": PatternFill(start_color=self.COLORS["header"], end_color=self.COLORS["header"], fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            ),
        }

    def _get_severity_fill(self, severity: str) -> PatternFill:
        """Get fill color based on severity."""
        color = self.COLORS.get(severity.lower(), "FFFFFF")
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def _apply_style(self, cell, **kwargs):
        """Apply style to cell."""
        if "font" in kwargs:
            cell.font = kwargs["font"]
        if "fill" in kwargs:
            cell.fill = kwargs["fill"]
        if "alignment" in kwargs:
            cell.alignment = kwargs["alignment"]
        if "border" in kwargs:
            cell.border = kwargs["border"]

    def _create_summary_sheet(self, metadata: Optional[Dict], findings: List[Dict]):
        """Create summary/overview sheet."""
        ws = self.workbook.create_sheet("Summary", 0)
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 50

        # Title
        ws["A1"] = "VAPT Scan Report - Summary"
        ws["A1"].font = Font(bold=True, size=16, color=self.COLORS["header"])
        ws.merge_cells("A1:B1")

        row = 3
        
        if metadata:
            # Metadata section
            ws[f"A{row}"] = "Scan Information"
            ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
            ws[f"A{row}"].fill = PatternFill(start_color=self.COLORS["header"], end_color=self.COLORS["header"], fill_type="solid")
            ws.merge_cells(f"A{row}:B{row}")
            row += 1

            info_items = [
                ("Target", metadata.get("target", "")),
                ("Scan Date", metadata.get("scan_date", "")),
                ("Scan Type", metadata.get("scan_type", "")),
                ("Total Findings", str(metadata.get("total_findings", 0))),
                ("Duration", metadata.get("duration", "")),
            ]

            for label, value in info_items:
                ws[f"A{row}"] = label
                ws[f"A{row}"].font = Font(bold=True)
                ws[f"B{row}"] = value
                row += 1

            row += 1

            # Severity Summary
            ws[f"A{row}"] = "Findings by Severity"
            ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
            ws[f"A{row}"].fill = PatternFill(start_color=self.COLORS["header"], end_color=self.COLORS["header"], fill_type="solid")
            ws.merge_cells(f"A{row}:B{row}")
            row += 1

            severity_counts = metadata.get("findings_by_severity", {})
            for severity in ["critical", "high", "medium", "low", "info"]:
                count = severity_counts.get(severity, 0)
                ws[f"A{row}"] = severity.capitalize()
                ws[f"A{row}"].font = Font(bold=True)
                ws[f"B{row}"] = count
                ws[f"B{row}"].fill = self._get_severity_fill(severity)
                ws[f"B{row}"].font = Font(bold=True, color="FFFFFF")
                ws[f"B{row}"].alignment = Alignment(horizontal="center")
                row += 1
        else:
            # Simple summary
            ws[f"A{row}"] = "Total Findings"
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = len(findings)
            row += 1

        row += 2

        # Finding type summary
        ws[f"A{row}"] = "Findings by Type"
        ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill(start_color=self.COLORS["header"], end_color=self.COLORS["header"], fill_type="solid")
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

        type_counts = {}
        for finding in findings:
            ftype = finding.get("type", "unknown")
            type_counts[ftype] = type_counts.get(ftype, 0) + 1

        for ftype, count in sorted(type_counts.items()):
            ws[f"A{row}"] = ftype.replace("_", " ").title()
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = count
            ws[f"B{row}"].alignment = Alignment(horizontal="center")
            row += 1

    def _create_findings_sheet(self, findings: List[Dict], include_evidence: bool):
        """Create detailed findings sheet."""
        ws = self.workbook.create_sheet("Findings")

        # Headers
        headers = [
            "ID",
            "Type",
            "Title",
            "Severity",
            "Confidence",
            "Location",
            "Description",
            "Remediation",
            "CWE ID",
            "OWASP",
        ]

        if include_evidence:
            headers.insert(7, "Evidence")

        header_style = self._get_header_style()

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            self._apply_style(cell, **header_style)

        # Set column widths
        column_widths = {
            1: 5,    # ID
            2: 15,   # Type
            3: 25,   # Title
            4: 12,   # Severity
            5: 12,   # Confidence
            6: 20,   # Location
            7: 30,   # Description
            8: 25,   # Remediation
            9: 10,   # CWE ID
            10: 20,  # OWASP
        }

        if include_evidence:
            column_widths[8] = 30  # Evidence

        for col, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        # Add findings
        for row_idx, finding in enumerate(findings, 2):
            col_idx = 1

            # ID
            ws.cell(row=row_idx, column=col_idx).value = row_idx - 1
            col_idx += 1

            # Type
            ws.cell(row=row_idx, column=col_idx).value = finding.get("type", "")
            col_idx += 1

            # Title
            ws.cell(row=row_idx, column=col_idx).value = finding.get("title", "")
            col_idx += 1

            # Severity
            severity_cell = ws.cell(row=row_idx, column=col_idx)
            severity_cell.value = finding.get("severity", "").upper()
            severity_cell.fill = self._get_severity_fill(finding.get("severity", ""))
            severity_cell.alignment = Alignment(horizontal="center")
            col_idx += 1

            # Confidence
            ws.cell(row=row_idx, column=col_idx).value = finding.get("confidence", "").capitalize()
            col_idx += 1

            # Location
            ws.cell(row=row_idx, column=col_idx).value = finding.get("location", "")
            col_idx += 1

            # Description
            desc_cell = ws.cell(row=row_idx, column=col_idx)
            desc_cell.value = finding.get("description", "")
            desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
            col_idx += 1

            # Evidence (if included)
            if include_evidence:
                evidence_cell = ws.cell(row=row_idx, column=col_idx)
                evidence_cell.value = finding.get("evidence", "")
                evidence_cell.alignment = Alignment(wrap_text=True, vertical="top")
                col_idx += 1

            # Remediation
            remediation_cell = ws.cell(row=row_idx, column=col_idx)
            remediation_cell.value = finding.get("remediation", "")
            remediation_cell.alignment = Alignment(wrap_text=True, vertical="top")
            col_idx += 1

            # CWE ID
            if finding.get("cwe_id"):
                ws.cell(row=row_idx, column=col_idx).value = f"CWE-{finding.get('cwe_id')}"
                ws.cell(row=row_idx, column=col_idx).font = Font(color="0563C1", underline="single")
            col_idx += 1

            # OWASP
            ws.cell(row=row_idx, column=col_idx).value = finding.get("owasp", "")
            col_idx += 1

        # Freeze panes
        ws.freeze_panes = "A2"

    def _create_evidence_sheet(self, findings: List[Dict]):
        """Create evidence sheet for findings with evidence."""
        ws = self.workbook.create_sheet("Evidence")
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 80

        # Headers
        for col_idx, header in enumerate(["Finding ID", "Title", "Evidence"], 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            self._apply_style(cell, **self._get_header_style())

        row = 2
        for idx, finding in enumerate(findings, 1):
            if finding.get("evidence"):
                ws.cell(row=row, column=1).value = idx
                ws.cell(row=row, column=2).value = finding.get("title", "")
                
                evidence_cell = ws.cell(row=row, column=3)
                evidence_cell.value = finding.get("evidence", "")
                evidence_cell.alignment = Alignment(wrap_text=True, vertical="top")
                
                row += 1

    def _create_timeline_sheet(self, findings: List[Dict]):
        """Create timeline sheet showing findings discovery."""
        ws = self.workbook.create_sheet("Timeline")
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 12

        # Headers
        for col_idx, header in enumerate(["Index", "Type", "Title", "Severity"], 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            self._apply_style(cell, **self._get_header_style())

        # Add findings in order
        for row_idx, finding in enumerate(findings, 2):
            ws.cell(row=row_idx, column=1).value = row_idx - 1
            ws.cell(row=row_idx, column=2).value = finding.get("type", "")
            ws.cell(row=row_idx, column=3).value = finding.get("title", "")
            
            severity_cell = ws.cell(row=row_idx, column=4)
            severity_cell.value = finding.get("severity", "").upper()
            severity_cell.fill = self._get_severity_fill(finding.get("severity", ""))
            severity_cell.alignment = Alignment(horizontal="center")

    def _create_statistics_sheet(self, findings: List[Dict], metadata: Optional[Dict]):
        """Create statistics sheet with analysis."""
        ws = self.workbook.create_sheet("Statistics")
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15

        row = 1

        # Severity Statistics
        ws[f"A{row}"] = "Severity Distribution"
        ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill(start_color=self.COLORS["header"], end_color=self.COLORS["header"], fill_type="solid")
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

        severity_counts = {}
        for finding in findings:
            severity = finding.get("severity", "info").lower()
            severity_counts[severity] = severity_counts.get(severity, 0) + 1

        for severity in ["critical", "high", "medium", "low", "info"]:
            count = severity_counts.get(severity, 0)
            ws[f"A{row}"] = severity.capitalize()
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = count
            ws[f"B{row}"].fill = self._get_severity_fill(severity)
            ws[f"B{row}"].alignment = Alignment(horizontal="center")
            row += 1

        row += 2

        # Type Statistics
        ws[f"A{row}"] = "Type Distribution"
        ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill(start_color=self.COLORS["header"], end_color=self.COLORS["header"], fill_type="solid")
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

        type_counts = {}
        for finding in findings:
            ftype = finding.get("type", "unknown")
            type_counts[ftype] = type_counts.get(ftype, 0) + 1

        for ftype, count in sorted(type_counts.items(), key=lambda x: x[1], reverse=True):
            ws[f"A{row}"] = ftype.replace("_", " ").title()
            ws[f"B{row}"] = count
            ws[f"B{row}"].alignment = Alignment(horizontal="center")
            row += 1

        row += 2

        # Confidence Statistics
        ws[f"A{row}"] = "Confidence Distribution"
        ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill(start_color=self.COLORS["header"], end_color=self.COLORS["header"], fill_type="solid")
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

        confidence_counts = {}
        for finding in findings:
            confidence = finding.get("confidence", "unknown").lower()
            confidence_counts[confidence] = confidence_counts.get(confidence, 0) + 1

        for confidence in ["high", "medium", "low"]:
            count = confidence_counts.get(confidence, 0)
            if count > 0:
                ws[f"A{row}"] = confidence.capitalize()
                ws[f"B{row}"] = count
                ws[f"B{row}"].alignment = Alignment(horizontal="center")
                row += 1
